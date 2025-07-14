
import os, tempfile
from flask import Flask, request, jsonify
from google.cloud import vision
import openpyxl
import smtplib
from email.message import EmailMessage

app = Flask(__name__)

# ---- Configuration ----
# These should be provided as environment variables on Render:
# GMAIL_USER  -> your sender Gmail address (e.g. pablo.al3435@gmail.com)
# GMAIL_PASS  -> 16‑znakowe hasło aplikacji wygenerowane w ustawieniach Google
# GOOGLE_APPLICATION_CREDENTIALS is automatically set to the path of the credentials JSON
GMAIL_USER = os.environ.get("GMAIL_USER")
GMAIL_PASS = os.environ.get("GMAIL_PASS")

if not GMAIL_USER or not GMAIL_PASS:
    raise RuntimeError("Brak zmiennych środowiskowych GMAIL_USER / GMAIL_PASS")

# Inicjalizacja klienta Vision AI używa pliku JSON określonego
# przez GOOGLE_APPLICATION_CREDENTIALS
client = vision.ImageAnnotatorClient()

@app.route('/upload', methods=['POST'])
def upload():
    """Przyjmij zdjęcie (multipart/form‑data 'file') + opcjonalne pole 'to' (adres odbiorcy)."""
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pola file w zapytaniu'}), 400

    # Odczytaj obraz
    img_file = request.files['file']
    image_content = img_file.read()

    # --- OCR ---
    image = vision.Image(content=image_content)
    response = client.text_detection(image=image)
    if response.error.message:
        return jsonify({'error': response.error.message}), 500

    raw_text = response.full_text_annotation.text

    # --- Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR"
    ws['A1'] = "Rozpoznany tekst:"
    ws['A2'] = raw_text

    tmp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_excel.name)

    # --- Email ---
    msg = EmailMessage()
    msg['Subject'] = 'Wyniki OCR'
    msg['From'] = GMAIL_USER
    recipient = request.form.get('to', GMAIL_USER)
    msg['To'] = recipient
    msg.set_content("W załączniku znajdziesz plik Excel z wynikami OCR.")

    with open(tmp_excel.name, 'rb') as fp:
        msg.add_attachment(fp.read(),
                           maintype='application',
                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           filename='wyniki.xlsx')

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(GMAIL_USER, GMAIL_PASS)
        smtp.send_message(msg)

    return jsonify({'status': 'ok', 'text': raw_text})

if __name__ == '__main__':
    # Render ustawia zmienną PORT – użyj jej
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
